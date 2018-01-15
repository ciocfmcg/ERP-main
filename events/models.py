from __future__ import unicode_literals

from django.db import models
from django.contrib.auth.models import User
from ERP.models import address
from time import time
from django.http import HttpResponse
# Create your models here.
from django.contrib import admin

def getEventsBannerUploadsPath(instance , filename ):
    return 'events/doc/%s_%s_%s' % (str(time()).replace('.', '_'), instance.contactPerson.username, filename)



def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Subscriptions"

    row_num = 0

    columns = [
        (u"ID", 8),
        (u"Name", 50),
        (u"Phone", 50),
        (u"Email", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            row_num,
            obj.name,
            obj.phone,
            obj.email,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            # c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"

class SubscriptionsAdmin(admin.ModelAdmin):
    actions = [export_xlsx]

class Event(models.Model):
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)
    name = models.CharField(max_length = 100 , null = False)
    dated = models.DateTimeField(null = False)
    contactPerson = models.ForeignKey(User, related_name='eventsManaging')
    venue = models.ForeignKey(address , null = False)
    bannerImage = models.FileField(upload_to = getEventsBannerUploadsPath , null = True ) # can be image , video or document
    tagLine = models.CharField(max_length = 100 , null = True)
    htmlDescription = models.TextField(max_length= 1000, null = True)
    active = models.BooleanField(default = False)
    def __str__(self):
        return "%s : %s" %(self.name , self.dated)

class Subscription(models.Model):
    created = models.DateTimeField(auto_now_add=True)
    active = models.BooleanField(default = True)
    name = models.CharField(max_length = 100 , null = False)
    email = models.EmailField(null = False)
    phone = models.PositiveIntegerField(null = False)
    comments = models.TextField(null = True)
    event = models.ForeignKey(Event , null = True, related_name='subscriptions')
    def __str__(self):
        return "%s : %s : %s" %(self.name , self.phone , self.email)

admin.site.register(Event)
admin.site.register(Subscription , SubscriptionsAdmin)
admin.site.register(address)
